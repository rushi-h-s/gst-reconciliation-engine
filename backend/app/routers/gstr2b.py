import json
from datetime import date, datetime
from io import BytesIO
from typing import Any, Callable

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from supabase import AsyncClient

from app.dependencies import get_org_id, get_supabase
from app.normalization import normalize_gstin, normalize_invoice_number

router = APIRouter()


# ── Endpoints ──────────────────────────────────────────────────────────────────

@router.post("/upload", status_code=201)
async def upload_gstr2b(
    file: UploadFile = File(...),
    client_id: str | None = None,
    period: str | None = None,
    org_id: str = Depends(get_org_id),
    supabase: AsyncClient = Depends(get_supabase),
):
    """Accept a GSTR-2B JSON file, parse all sections, and insert entries."""
    content_type = file.content_type or ""
    if "json" not in content_type and not (file.filename or "").endswith(".json"):
        raise HTTPException(status_code=415, detail="Only JSON GSTR-2B uploads supported on this endpoint")

    raw = await file.read()
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {exc}")

    by_section, parse_warnings = _parse_gstr2b_json(payload, org_id, client_id or "", period or "")
    all_entries = [e for entries in by_section.values() for e in entries]

    if not all_entries:
        return {"inserted": 0, "by_section": {}, "parse_warnings": parse_warnings,
                "warning": "No entries found in the uploaded file"}

    await supabase.table("gstr2b_entries").insert(all_entries).execute()
    return {
        "inserted": len(all_entries),
        "by_section": {k: len(v) for k, v in by_section.items() if v},
        "parse_warnings": parse_warnings,
    }


@router.post("/upload-excel", status_code=201)
async def upload_gstr2b_excel(
    file: UploadFile = File(...),
    client_id: str | None = None,
    period: str | None = None,
    org_id: str = Depends(get_org_id),
    supabase: AsyncClient = Depends(get_supabase),
):
    """Accept a GSTR-2B Excel (.xlsx) file from the GST portal.

    Parses B2B, B2BA, CDNR, IMPG, and ISDA sheets and inserts all found
    entries into gstr2b_entries.
    """
    filename = file.filename or ""
    content_type = file.content_type or ""
    is_xlsx = filename.endswith(".xlsx") or "spreadsheet" in content_type or "excel" in content_type
    if not is_xlsx:
        raise HTTPException(status_code=415, detail="Only .xlsx files are accepted on this endpoint")

    raw = await file.read()
    try:
        by_section, parse_warnings = _parse_gstr2b_excel(raw, org_id, client_id or "", period or "")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel parse error: {exc}")

    all_entries = [e for entries in by_section.values() for e in entries]

    if not all_entries:
        return {"inserted": 0, "by_section": {}, "parse_warnings": parse_warnings,
                "warning": "No entries found in the uploaded file"}

    await supabase.table("gstr2b_entries").insert(all_entries).execute()
    return {
        "inserted": len(all_entries),
        "by_section": {k: len(v) for k, v in by_section.items() if v},
        "parse_warnings": parse_warnings,
    }


# ── Shared helpers ────────────────────────────────────────────────────────────

_MONTH_ABBR: dict[str, str] = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}


def _parse_excel_date(value: object) -> str:
    """Normalise any date representation to YYYY-MM-DD."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    parts = s.replace("/", "-").split("-")
    if len(parts) == 3:
        d, m, y = parts
        if m.lower() in _MONTH_ABBR and len(y) == 4:
            return f"{y}-{_MONTH_ABBR[m.lower()]}-{d.zfill(2)}"
        if len(y) == 4 and m.isdigit():
            return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return s


def _to_amt(value: object) -> str:
    try:
        return str(float(value or 0))
    except (TypeError, ValueError):
        return "0"


def _find_sheet(
    wb: openpyxl.Workbook,
    keyword: str,
    exclude: str | None = None,
) -> Any | None:
    """Return the first sheet whose name matches *keyword* (case-insensitive).

    Tries exact match, then prefix, then substring. Sheets whose names
    contain *exclude* are skipped.
    """
    kw = keyword.lower()
    ex = exclude.lower() if exclude else None

    def _ok(name: str) -> bool:
        nl = name.lower()
        return ex is None or ex not in nl

    for name in wb.sheetnames:
        if name.lower() == kw and _ok(name):
            return wb[name]
    for name in wb.sheetnames:
        if name.lower().startswith(kw) and _ok(name):
            return wb[name]
    for name in wb.sheetnames:
        if kw in name.lower() and _ok(name):
            return wb[name]
    return None


def _extract_rows(
    ws: Any,
    map_fn: Callable[[object], str | None],
    max_scan: int = 15,
) -> list[dict[str, object]]:
    """Detect the header row and return data as a list of field-keyed dicts.

    Returns [] if no recognisable header is found (sheet is silently skipped).
    When two columns map to the same field the one with the higher column index
    wins — callers exploit this to prefer "Revised …" over "Original …" columns
    by putting them later in the sheet (as the GST portal does).
    """
    col_map: dict[int, str] = {}
    header_row_idx: int | None = None

    for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan), start=1):
        candidate: dict[int, str] = {}
        for col_idx, cell in enumerate(row):
            field = map_fn(cell.value)
            if field:
                candidate[col_idx] = field
        if len(candidate) >= 2:
            col_map = candidate
            header_row_idx = row_idx
            break

    if not col_map or header_row_idx is None:
        return []

    rows: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        cells = list(row)
        row_data: dict[str, object] = {
            field: cells[col_idx].value
            for col_idx, field in col_map.items()
            if col_idx < len(cells)
        }
        rows.append(row_data)
    return rows


def _make_entry(
    org_id: str,
    client_id: str,
    period: str,
    raw_gstin: str,
    name: str | None,
    inv_no: str,
    inv_date: str,
    taxable_value: object,
    cgst: object,
    sgst: object,
    igst: object,
    doc_type: str,
) -> dict:
    return {
        "org_id":              org_id,
        "client_id":           client_id,
        "period":              period,
        "supplier_gstin":      raw_gstin,
        "supplier_name":       name or None,
        "inv_no":              inv_no,
        "inv_date":            inv_date,
        "taxable_value":       _to_amt(taxable_value),
        "cgst":                _to_amt(cgst),
        "sgst":                _to_amt(sgst),
        "igst":                _to_amt(igst),
        "doc_type":            doc_type,
        "norm_supplier_gstin": normalize_gstin(raw_gstin),
        "norm_inv_no":         normalize_invoice_number(inv_no),
    }


# ── B2B ───────────────────────────────────────────────────────────────────────

def _map_b2b_col(cell_value: object) -> str | None:
    if cell_value is None:
        return None
    h = str(cell_value).lower().strip()
    if "gstin" in h:
        return "supplier_gstin"
    if "trade" in h or ("legal" in h and "name" in h):
        return "supplier_name"
    if "invoice" in h and "date" in h:
        return "inv_date"
    if "invoice" in h and ("no" in h or "number" in h):
        return "inv_no"
    if "taxable" in h and "value" in h:
        return "taxable_value"
    if "integrated" in h or "igst" in h:
        return "igst"
    if "central" in h or "cgst" in h:
        return "cgst"
    if ("state" in h or "ut tax" in h or "sgst" in h) and ("tax" in h or "amount" in h):
        return "sgst"
    return None


def _parse_excel_b2b(ws: Any, org_id: str, client_id: str, period: str) -> list[dict]:
    entries = []
    for row_data in _extract_rows(ws, _map_b2b_col):
        raw_gstin = str(row_data.get("supplier_gstin") or "").strip()
        if not raw_gstin:
            continue
        raw_inv_no = str(row_data.get("inv_no") or "").strip()
        raw_date = row_data.get("inv_date")
        entries.append(_make_entry(
            org_id, client_id, period,
            raw_gstin,
            str(row_data.get("supplier_name") or "").strip() or None,
            raw_inv_no,
            _parse_excel_date(raw_date) if raw_date else "",
            row_data.get("taxable_value"),
            row_data.get("cgst"),
            row_data.get("sgst"),
            row_data.get("igst"),
            "invoice",
        ))
    return entries


# ── B2BA (Amended B2B) ────────────────────────────────────────────────────────

def _map_b2ba_col(cell_value: object) -> str | None:
    if cell_value is None:
        return None
    h = str(cell_value).lower().strip()
    # Skip original-invoice columns — we only want the revised (amended) values.
    if "original" in h:
        return None
    # Prefer "Revised Invoice …" labels, then fall back to plain B2B mapping.
    if "revised" in h and "invoice" in h and "date" in h:
        return "inv_date"
    if "revised" in h and "invoice" in h and ("no" in h or "number" in h):
        return "inv_no"
    return _map_b2b_col(cell_value)


def _parse_excel_b2ba(
    ws: Any, org_id: str, client_id: str, period: str
) -> tuple[list[dict], list[str]]:
    entries: list[dict] = []
    warnings: list[str] = []
    for row_data in _extract_rows(ws, _map_b2ba_col):
        raw_gstin = str(row_data.get("supplier_gstin") or "").strip()
        if not raw_gstin:
            continue
        # _map_b2ba_col drops "original" columns, so inv_no here is always the
        # revised number. If it is empty the row cannot be identified — skip it.
        raw_inv_no = str(row_data.get("inv_no") or "").strip()
        if not raw_inv_no:
            warnings.append(
                f"B2BA row skipped (supplier GSTIN {raw_gstin!r}): "
                "no revised or original invoice number found"
            )
            continue
        raw_date = row_data.get("inv_date")
        entries.append(_make_entry(
            org_id, client_id, period,
            raw_gstin,
            str(row_data.get("supplier_name") or "").strip() or None,
            raw_inv_no,
            _parse_excel_date(raw_date) if raw_date else "",
            row_data.get("taxable_value"),
            row_data.get("cgst"),
            row_data.get("sgst"),
            row_data.get("igst"),
            "invoice",
        ))
    return entries, warnings


# ── CDNR (Credit / Debit Notes Received) ──────────────────────────────────────

def _map_cdnr_col(cell_value: object) -> str | None:
    if cell_value is None:
        return None
    h = str(cell_value).lower().strip()
    if "gstin" in h:
        return "supplier_gstin"
    if "trade" in h or ("legal" in h and "name" in h):
        return "supplier_name"
    # Note-specific columns take priority over plain "invoice" columns.
    if "note" in h and "type" in h:
        return "note_type"
    if "note" in h and "date" in h:
        return "inv_date"
    if "note" in h and ("no" in h or "number" in h):
        return "inv_no"
    if "document" in h and "type" in h:
        return "note_type"
    if "document" in h and "date" in h:
        return "inv_date"
    if "document" in h and ("no" in h or "number" in h):
        return "inv_no"
    if "taxable" in h and "value" in h:
        return "taxable_value"
    if "integrated" in h or "igst" in h:
        return "igst"
    if "central" in h or "cgst" in h:
        return "cgst"
    if ("state" in h or "ut tax" in h or "sgst" in h) and ("tax" in h or "amount" in h):
        return "sgst"
    return None


def _parse_excel_cdnr(ws: Any, org_id: str, client_id: str, period: str) -> list[dict]:
    entries = []
    for row_data in _extract_rows(ws, _map_cdnr_col):
        raw_gstin = str(row_data.get("supplier_gstin") or "").strip()
        if not raw_gstin:
            continue
        nt = str(row_data.get("note_type") or "").lower()
        doc_type = "debit_note" if ("debit" in nt or nt.strip() == "d") else "credit_note"
        raw_inv_no = str(row_data.get("inv_no") or "").strip()
        raw_date = row_data.get("inv_date")
        entries.append(_make_entry(
            org_id, client_id, period,
            raw_gstin,
            str(row_data.get("supplier_name") or "").strip() or None,
            raw_inv_no,
            _parse_excel_date(raw_date) if raw_date else "",
            row_data.get("taxable_value"),
            row_data.get("cgst"),
            row_data.get("sgst"),
            row_data.get("igst"),
            doc_type,
        ))
    return entries


# ── IMPG (Import of Goods via Bill of Entry) ──────────────────────────────────

def _map_impg_col(cell_value: object) -> str | None:
    if cell_value is None:
        return None
    h = str(cell_value).lower().strip()
    if "port" in h and "code" in h:
        return "port_code"
    if "bill" in h and "entry" in h and "date" in h:
        return "inv_date"
    if "bill" in h and "entry" in h and ("no" in h or "number" in h):
        return "inv_no"
    if "taxable" in h and "value" in h:
        return "taxable_value"
    # Imports carry only IGST (inter-state by definition).
    if "integrated" in h or "igst" in h:
        return "igst"
    return None


def _parse_excel_impg(ws: Any, org_id: str, client_id: str, period: str) -> list[dict]:
    entries = []
    for row_data in _extract_rows(ws, _map_impg_col):
        raw_inv_no = str(row_data.get("inv_no") or "").strip()
        if not raw_inv_no:
            continue
        port_code = str(row_data.get("port_code") or "").strip().upper()
        # Imports have no supplier GSTIN; encode port code so it's identifiable.
        raw_gstin = f"IMPG-{port_code}" if port_code else "IMPG-UNKNOWN"
        raw_date = row_data.get("inv_date")
        entries.append(_make_entry(
            org_id, client_id, period,
            raw_gstin,
            None,  # no supplier name for imports
            raw_inv_no,
            _parse_excel_date(raw_date) if raw_date else "",
            row_data.get("taxable_value"),
            0,  # no CGST on imports
            0,  # no SGST on imports
            row_data.get("igst"),
            "invoice",
        ))
    return entries


# ── ISDA (ISD Amendments) ─────────────────────────────────────────────────────

def _map_isda_col(cell_value: object) -> str | None:
    if cell_value is None:
        return None
    h = str(cell_value).lower().strip()
    # Skip original-document columns.
    if "original" in h:
        return None
    # ISD-specific GSTIN / name labels take priority.
    if ("isd" in h or "distributor" in h) and "gstin" in h:
        return "supplier_gstin"
    if ("isd" in h or "distributor" in h) and "name" in h:
        return "supplier_name"
    # Revised document number / date.
    if "revised" in h and "document" in h and "date" in h:
        return "inv_date"
    if "revised" in h and "document" in h and ("no" in h or "number" in h):
        return "inv_no"
    # Fallback: plain document columns.
    if "document" in h and "date" in h:
        return "inv_date"
    if "document" in h and ("no" in h or "number" in h):
        return "inv_no"
    # Shared amount columns.
    if "taxable" in h and "value" in h:
        return "taxable_value"
    if "integrated" in h or "igst" in h:
        return "igst"
    if "central" in h or "cgst" in h:
        return "cgst"
    if ("state" in h or "ut tax" in h or "sgst" in h) and ("tax" in h or "amount" in h):
        return "sgst"
    # Generic GSTIN fallback (plain header with just "GSTIN").
    if "gstin" in h:
        return "supplier_gstin"
    return None


def _parse_excel_isda(ws: Any, org_id: str, client_id: str, period: str) -> list[dict]:
    entries = []
    for row_data in _extract_rows(ws, _map_isda_col):
        raw_gstin = str(row_data.get("supplier_gstin") or "").strip()
        if not raw_gstin:
            continue
        raw_inv_no = str(row_data.get("inv_no") or "").strip()
        raw_date = row_data.get("inv_date")
        entries.append(_make_entry(
            org_id, client_id, period,
            raw_gstin,
            str(row_data.get("supplier_name") or "").strip() or None,
            raw_inv_no,
            _parse_excel_date(raw_date) if raw_date else "",
            row_data.get("taxable_value"),
            row_data.get("cgst"),
            row_data.get("sgst"),
            row_data.get("igst"),
            "invoice",
        ))
    return entries


# ── Excel orchestrator ────────────────────────────────────────────────────────

_EXCEL_SECTIONS: list[tuple[str, str | None, Any]] = [
    # (sheet_keyword, exclude_keyword, parser_fn)
    ("b2b",  "b2ba", _parse_excel_b2b),
    ("b2ba", None,   _parse_excel_b2ba),
    ("cdnr", None,   _parse_excel_cdnr),
    ("cdn",  None,   _parse_excel_cdnr),   # some portal versions use "CDN"
    ("impg", None,   _parse_excel_impg),
    ("isda", None,   _parse_excel_isda),
]


def _parse_gstr2b_excel(
    raw: bytes, org_id: str, client_id: str, period: str
) -> tuple[dict[str, list[dict]], list[str]]:
    """Parse all GSTR-2B sections from a portal Excel workbook.

    Returns ({section_name: [entries]}, parse_warnings).
    Missing sections produce no key (not an empty list).
    """
    wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
    results: dict[str, list[dict]] = {}
    all_warnings: list[str] = []
    seen_sheets: set[str] = set()

    for keyword, exclude, parse_fn in _EXCEL_SECTIONS:
        ws = _find_sheet(wb, keyword, exclude=exclude)
        if ws is None or ws.title in seen_sheets:
            continue
        seen_sheets.add(ws.title)

        if keyword == "b2ba":
            entries, warnings = parse_fn(ws, org_id, client_id, period)
            all_warnings.extend(warnings)
        else:
            entries = parse_fn(ws, org_id, client_id, period)

        if entries:
            section_key = "cdnr" if keyword in ("cdn", "cdnr") else keyword
            results.setdefault(section_key, []).extend(entries)

    return results, all_warnings


# ── JSON parser ───────────────────────────────────────────────────────────────

def _parse_gstr2b_json(
    data: dict, org_id: str, client_id: str, period: str
) -> tuple[dict[str, list[dict]], list[str]]:
    """Parse all GSTR-2B sections from the government JSON format.

    Returns ({section_name: [entries]}, parse_warnings).
    """
    docdata: dict = data.get("data", {}).get("docdata", {})
    results: dict[str, list[dict]] = {}
    all_warnings: list[str] = []

    # ── B2B ───────────────────────────────────────────────────────────────────
    b2b_entries: list[dict] = []
    for supplier in docdata.get("b2b", []):
        raw_gstin = supplier.get("ctin", "")
        name = supplier.get("trdnm", "")
        for inv in supplier.get("inv", []):
            for item in (inv.get("items") or [inv]):
                b2b_entries.append(_make_entry(
                    org_id, client_id, period,
                    raw_gstin, name,
                    inv.get("inum", ""),
                    inv.get("dt", ""),
                    item.get("txval", 0),
                    item.get("camt", 0),
                    item.get("samt", 0),
                    item.get("iamt", 0),
                    "invoice",
                ))
    if b2b_entries:
        results["b2b"] = b2b_entries

    # ── B2BA (Amended B2B) ────────────────────────────────────────────────────
    b2ba_entries: list[dict] = []
    for supplier in docdata.get("b2ba", []):
        raw_gstin = supplier.get("ctin", "")
        name = supplier.get("trdnm", "")
        for inv in supplier.get("inv", []):
            inv_no = inv.get("inum") or inv.get("oinum", "")
            if not inv_no:
                all_warnings.append(
                    f"B2BA entry skipped (supplier {raw_gstin!r}): "
                    "inum and oinum both empty"
                )
                continue
            inv_date = inv.get("dt") or inv.get("oidt", "")
            for item in (inv.get("items") or [inv]):
                b2ba_entries.append(_make_entry(
                    org_id, client_id, period,
                    raw_gstin, name,
                    inv_no, inv_date,
                    item.get("txval", 0),
                    item.get("camt", 0),
                    item.get("samt", 0),
                    item.get("iamt", 0),
                    "invoice",
                ))
    if b2ba_entries:
        results["b2ba"] = b2ba_entries

    # ── CDNR (Credit / Debit Notes) ───────────────────────────────────────────
    cdnr_entries: list[dict] = []
    for supplier in docdata.get("cdnr", []):
        raw_gstin = supplier.get("ctin", "")
        name = supplier.get("trdnm", "")
        for note in supplier.get("nt", []):
            typ = str(note.get("typ") or "C").upper()
            doc_type = "debit_note" if typ == "D" else "credit_note"
            for item in (note.get("items") or [note]):
                cdnr_entries.append(_make_entry(
                    org_id, client_id, period,
                    raw_gstin, name,
                    note.get("ntnum", ""),
                    note.get("dt", ""),
                    item.get("txval", 0),
                    item.get("camt", 0),
                    item.get("samt", 0),
                    item.get("iamt", 0),
                    doc_type,
                ))
    if cdnr_entries:
        results["cdnr"] = cdnr_entries

    # ── IMPG (Import of Goods) ────────────────────────────────────────────────
    impg_entries: list[dict] = []
    for imp in docdata.get("impg", []):
        port_code = str(imp.get("portcd") or "").strip().upper()
        raw_gstin = f"IMPG-{port_code}" if port_code else "IMPG-UNKNOWN"
        impg_entries.append(_make_entry(
            org_id, client_id, period,
            raw_gstin, None,
            imp.get("boenum", ""),
            imp.get("bodt", ""),
            imp.get("txval", 0),
            0,             # no CGST on imports
            0,             # no SGST on imports
            imp.get("iamt", 0),
            "invoice",
        ))
    if impg_entries:
        results["impg"] = impg_entries

    # ── ISDA (ISD Amendments) ─────────────────────────────────────────────────
    isda_entries: list[dict] = []
    for distributor in docdata.get("isda", []):
        raw_gstin = distributor.get("ctin", "")
        name = distributor.get("trdnm", "")
        for doc in distributor.get("doclist", []):
            doc_no = doc.get("docnum") or doc.get("odocnum", "")
            doc_date = doc.get("docdt") or doc.get("odocdt", "")
            for item in (doc.get("items") or [doc]):
                isda_entries.append(_make_entry(
                    org_id, client_id, period,
                    raw_gstin, name,
                    doc_no, doc_date,
                    item.get("txval", 0),
                    item.get("camt", 0),
                    item.get("samt", 0),
                    item.get("iamt", 0),
                    "invoice",
                ))
    if isda_entries:
        results["isda"] = isda_entries

    return results, all_warnings
